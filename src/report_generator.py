"""
Builds the daily performance report: an Excel workbook (summary + charts)
and a companion HTML version, both saved into reports/.
"""
import logging
from datetime import datetime
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import config

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
ALERT_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")


def _write_df_sheet(ws, df: pd.DataFrame, highlight_col: str = None, highlight_below: float = None):
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    if highlight_col and highlight_col in df.columns:
        col_idx = list(df.columns).index(highlight_col) + 1
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=col_idx).value
            if val is not None and val < highlight_below:
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = ALERT_FILL
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 30)


def _efficiency_chart(panel_summary: pd.DataFrame, out_path: Path) -> Path:
    fig, ax = plt.subplots(figsize=(9, 4))
    worst = panel_summary.sort_values("avg_efficiency").head(20)
    ax.bar(worst["panel_id"].astype(str), worst["avg_efficiency"] * 100, color="#c0504d")
    ax.axhline(config.UNDERPERFORMANCE_THRESHOLD * 100, color="black", linestyle="--",
               label=f"{config.UNDERPERFORMANCE_THRESHOLD*100:.0f}% threshold")
    ax.set_ylabel("Avg efficiency (%)")
    ax.set_title("20 Lowest-Efficiency Panels")
    ax.set_xticks(range(len(worst)))
    ax.set_xticklabels(worst["panel_id"].astype(str), rotation=90, fontsize=6)
    ax.legend()
    fig.tight_layout()
    fig.savefig(out_path, dpi=150)
    plt.close(fig)
    return out_path


def generate_daily_report(readings: pd.DataFrame, panel_summary: pd.DataFrame,
                           underperforming: pd.DataFrame,
                           report_dir: Path = config.REPORT_DIR) -> dict:
    """Write an Excel + HTML daily report. Returns dict of output paths."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = report_dir / f"daily_report_{ts}.xlsx"
    html_path = report_dir / f"daily_report_{ts}.html"
    chart_path = report_dir / f"efficiency_chart_{ts}.png"

    total_panels = readings["panel_id"].nunique()
    n_underperforming = underperforming["panel_id"].nunique()
    plant_avg_efficiency = readings["efficiency"].mean(skipna=True)
    total_power_kw = readings.groupby("panel_id")["power_w"].mean().sum() / 1000

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Solar Plant Daily Performance Report"])
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.append(["Generated", datetime.now().isoformat(timespec="seconds")])
    ws_summary.append(["Total panels monitored", total_panels])
    ws_summary.append(["Panels underperforming (>=20% below expected)", n_underperforming])
    ws_summary.append(["Plant average efficiency", f"{plant_avg_efficiency*100:.1f}%" if pd.notna(plant_avg_efficiency) else "n/a"])
    ws_summary.append(["Estimated current plant output (kW)", round(total_power_kw, 1)])

    if not panel_summary.empty:
        _efficiency_chart(panel_summary, chart_path)
        img = XLImage(str(chart_path))
        ws_summary.add_image(img, "A9")

    ws_panels = wb.create_sheet("Panel Summary")
    _write_df_sheet(ws_panels, panel_summary.round(3), highlight_col="avg_efficiency",
                     highlight_below=config.UNDERPERFORMANCE_THRESHOLD)

    ws_alerts = wb.create_sheet("Underperforming Panels")
    if not underperforming.empty:
        cols = ["timestamp", "panel_id", "power_w", "expected_power_w", "efficiency", "shortfall_pct"]
        _write_df_sheet(ws_alerts, underperforming[cols].round(3))
    else:
        ws_alerts.append(["No underperforming panels detected in this batch."])

    wb.save(xlsx_path)

    html = f"""
    <html><head><meta charset='utf-8'><title>Solar Plant Daily Report</title>
    <style>
      body {{ font-family: Arial, sans-serif; margin: 2rem; color: #1a1a1a; }}
      h1 {{ color: #1F4E78; }}
      table {{ border-collapse: collapse; width: 100%; margin-top: 1rem; }}
      th, td {{ border: 1px solid #ddd; padding: 6px 10px; text-align: left; font-size: 13px; }}
      th {{ background: #1F4E78; color: white; }}
      .alert {{ background: #F8CBAD; }}
      .metric {{ font-size: 1.1rem; margin: 4px 0; }}
    </style></head><body>
    <h1>Solar Plant Daily Performance Report</h1>
    <p class='metric'><b>Generated:</b> {datetime.now().isoformat(timespec='seconds')}</p>
    <p class='metric'><b>Total panels monitored:</b> {total_panels}</p>
    <p class='metric'><b>Panels underperforming:</b> {n_underperforming}</p>
    <p class='metric'><b>Plant average efficiency:</b> {plant_avg_efficiency*100:.1f}%</p>
    <p class='metric'><b>Estimated current output:</b> {total_power_kw:.1f} kW</p>
    <img src='{chart_path.name}' style='max-width:800px'/>
    <h2>Underperforming Panels</h2>
    {underperforming[['timestamp','panel_id','power_w','expected_power_w','efficiency','shortfall_pct']].round(2).to_html(index=False) if not underperforming.empty else '<p>None in this batch.</p>'}
    </body></html>
    """
    html_path.write_text(html, encoding="utf-8")

    logger.info("report_generator: wrote %s and %s", xlsx_path.name, html_path.name)
    return {"xlsx": xlsx_path, "html": html_path, "chart": chart_path}
